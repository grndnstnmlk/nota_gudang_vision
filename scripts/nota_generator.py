#!/usr/bin/env python3
"""Generator NOTA PEMBELIAN TEMBAKAU (xlsx 2) otomatis dari Buku Sortir (xlsx 1)."""
import re
from dataclasses import dataclass
from typing import Optional
import os
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

COL = {"GL":1,"NO":2,"GT":3,"NAMA":4,"GRADE":5,"HARGA":6,"KG":7,"BRT":8,"BRT_FIX":9,"NET":10,"KET":11}
# Blok "BS" di xlsx 1: Penanda "bs" di kolom L (dan KET kolom K biasanya berisi "ada bs").
BS = {"MARK":12,"KETMARK":11,"GRADE":14,"HARGA":15,"KG":16,"BRT":17,"NET":18}
import math

def _round_int(x):
    """Bulatkan ke bilangan bulat terdekat (half-up). Kembalikan apa adanya bila bukan angka."""
    if x is None or x=="" : return x
    try: xf=float(x)
    except (TypeError,ValueError): return x
    return int(math.floor(xf+0.5)) if xf>=0 else -int(math.floor(-xf+0.5))

def calc_brt(kg, brt_fix=None):
    """Hitung BRT: jika brt_fix terisi gunakan brt_fix; jika tidak, gunakan aturan pecahan KG."""
    if brt_fix not in (None, ""):
        return _round_int(brt_fix)
    if kg is None or kg == "": return None
    try: kf = float(kg)
    except (TypeError, ValueError): return kg
    dec = round(kf % 1, 1)
    if dec > 0 and dec <= 0.4:
        return int(kf) - 1
    return int(kf)

def calc_harga(grade, harga=None):
    """Hitung HARGA: GRADE * 1000 jika harga belum dievaluasi/formula."""
    if harga not in (None, "") and not (isinstance(harga, str) and harga.startswith("=")):
        try: return _round_int(harga)
        except (TypeError, ValueError): pass
    if grade not in (None, ""):
        try: return _round_int(float(grade) * 1000)
        except (TypeError, ValueError): pass
    return harga

def calc_net(brt, gl=None):
    """Hitung NET berdasarkan aturan: GL -> BRT-2, BRT>=60 -> BRT-5, BRT>=50 -> BRT-4, BRT>=10 -> BRT-3."""
    if brt is None or brt == "": return None
    try: b = float(brt)
    except (TypeError, ValueError): return brt
    if str(gl or "").strip().lower() == "gl": return _round_int(b - 2)
    elif b >= 60: return _round_int(b - 5)
    elif b >= 50: return _round_int(b - 4)
    elif b >= 10: return _round_int(b - 3)
    else: return _round_int(b)

BULAN = ["","Januari","Februari","Maret","April","Mei","Juni","Juli","Agustus",
         "September","Oktober","November","Desember"]
DATE_RE = re.compile(r"^\(?\s*(\d{1,2})\s*[/\-.]\s*(\d{1,2})\s*[/\-.]\s*(\d{2,4})\s*\)?$")
# tanggal berbentuk teks Indonesia, mis. "10 Agustus 2026" atau "1 mei 26"
_BULAN_IDX = {b.lower(): i for i, b in enumerate(BULAN) if b}
TEXT_DATE_RE = re.compile(
    r"^\(?\s*(\d{1,2})\s+([A-Za-z]+)\s+(\d{2,4})\s*\)?$")

def _row_of(no): return no + 3

def _is_date_token(v):
    if not v: return False
    s = str(v).strip()
    if DATE_RE.match(s): return True
    m = TEXT_DATE_RE.match(s)
    return bool(m) and m.group(2).lower() in _BULAN_IDX

def _fmt_indo_date(v):
    if v is None: return ""
    s = str(v).strip()
    m = DATE_RE.match(s)
    if m:
        d,mth,y = int(m.group(1)),int(m.group(2)),int(m.group(3))
        if y<100: y+=2000
        mth=min(max(mth,1),12)
        return f"{d} {BULAN[mth]} {y}"
    m = TEXT_DATE_RE.match(s)
    if m and m.group(2).lower() in _BULAN_IDX:
        d = int(m.group(1)); mth = _BULAN_IDX[m.group(2).lower()]; y = int(m.group(3))
        if y<100: y+=2000
        return f"{d} {BULAN[mth]} {y}"
    return s

def _has_bs(ws, r):
    """True bila baris r punya penanda BS (kolom L='bs' atau KET kolom J memuat 'bs')."""
    mk=ws.cell(r,BS["MARK"]).value
    kt=ws.cell(r,BS["KETMARK"]).value
    if isinstance(mk,str) and mk.strip().lower()=="bs": return True
    if isinstance(kt,str) and "bs" in kt.strip().lower(): return True
    return False

def read_rows(ws, a, b):
    out=[]; bs_rows=[]
    for no in range(a,b+1):
        r=_row_of(no)
        # --- baris utama (GL) ---
        gl=ws.cell(r,COL["GL"]).value
        gt=ws.cell(r,COL["GT"]).value
        grade=ws.cell(r,COL["GRADE"]).value
        kg=ws.cell(r,COL["KG"]).value
        brt_fix=ws.cell(r,COL.get("BRT_FIX",9)).value if "BRT_FIX" in COL else None
        brt_raw=ws.cell(r,COL["BRT"]).value
        net_raw=ws.cell(r,COL["NET"]).value
        harga_raw=ws.cell(r,COL["HARGA"]).value

        # Ambil harga (hitung dari grade bila formula/None)
        harga = calc_harga(grade, harga_raw)

        # Ambil BRT (prioritas BRT FIX, lalu nilai terisi, lalu hitung dari KG)
        if brt_fix not in (None, ""):
            brt = _round_int(brt_fix)
        elif brt_raw not in (None, "") and not (isinstance(brt_raw, str) and brt_raw.startswith("=")):
            brt = _round_int(brt_raw)
        else:
            brt = calc_brt(kg)

        # Ambil NET (hitung jika formula/None)
        if net_raw not in (None, "") and not (isinstance(net_raw, str) and net_raw.startswith("=")):
            net = _round_int(net_raw)
        else:
            net = calc_net(brt, gl)

        if not (brt is None and net is None and harga is None):
            out.append({"no":no,"gl":gl,"gt":gt,"bruto":brt,"netto":net,"harga":harga,"bs":False})
        # --- blok BS di kanan (kalau ada) ---
        if _has_bs(ws,r):
            bh=ws.cell(r,BS["HARGA"]).value; bb=ws.cell(r,BS["BRT"]).value; bn=ws.cell(r,BS["NET"]).value
            bs_grd=ws.cell(r,BS["GRADE"]).value
            bs_kg=ws.cell(r,BS["KG"]).value
            b_harga=calc_harga(bs_grd, bh)
            b_brt=_round_int(bb) if bb not in (None, "") and not (isinstance(bb, str) and bb.startswith("=")) else calc_brt(bs_kg)
            b_net=_round_int(bn) if bn not in (None, "") and not (isinstance(bn, str) and bn.startswith("=")) else calc_net(b_brt)
            if not (b_harga is None and b_brt is None and b_net is None):
                bs_rows.append({"no":no,"gt":None,"bruto":b_brt,
                                "netto":b_net,"harga":b_harga,"bs":True,"label":"BS"})
    # baris BS selalu diletakkan setelah semua baris GL
    out.extend(bs_rows)
    return out

def detect_info(ws,a,b):
    """Baca kolom NAMA untuk rentang NO a..b. Susunannya per orang (dari atas ke bawah):
        baris-1 = NAMA, baris-2 = TANGGAL, baris-3 = ALAMAT.
    Kembalikan (nama, tanggal, alamat). Tanggal dipakai sebagai patokan:
    teks tepat SEBELUM tanggal = nama, teks SETELAH tanggal = alamat."""
    items=[]  # (no, teks, is_date)
    for no in range(a,b+1):
        v=ws.cell(_row_of(no),COL["NAMA"]).value
        if v in (None,""): continue
        items.append((no, str(v).strip(), _is_date_token(v)))
    # cari token tanggal pertama sebagai patokan
    di = next((i for i,(_,_,d) in enumerate(items) if d), None)
    nama=""; tanggal=""; alamat=""
    if di is not None:
        tanggal=_fmt_indo_date(items[di][1])
        # nama = teks terakhir sebelum tanggal
        for i in range(di-1,-1,-1):
            if not items[i][2]: nama=items[i][1]; break
        # alamat = teks pertama sesudah tanggal
        for i in range(di+1,len(items)):
            if not items[i][2]: alamat=items[i][1]; break
    else:
        # tidak ada tanggal di rentang -> teks pertama dianggap nama, kedua alamat
        texts=[t for _,t,d in items if not d]
        if texts: nama=texts[0]
        if len(texts)>1: alamat=texts[1]
    # kalau nama tetap kosong (rentang mulai di bawah nama), cari ke atas
    if not nama:
        for no in range(a-1,0,-1):
            v=ws.cell(_row_of(no),COL["NAMA"]).value
            if v and not _is_date_token(v): nama=str(v).strip(); break
    return nama, tanggal, alamat

FONT_TITLE=dict(name="Bahnschrift",size=16,bold=True)
FONT_LABEL=dict(name="Bahnschrift",size=11,bold=True)
FONT_VALUE=dict(name="Bahnschrift",size=11,bold=False)
FONT_HEAD =dict(name="Bahnschrift",size=12,bold=True)
FONT_GL   =dict(name="Bahnschrift",size=11,bold=False)
FONT_CELL =dict(name="Bahnschrift",size=12,bold=False)
FONT_TOTAL=dict(name="Bahnschrift",size=12,bold=True)
THIN=Side(style="thin"); MED=Side(style="medium"); NONE=Side(style=None)
COL_WIDTH={"A":13.28515625,"B":6.140625,"C":12.85546875,"D":13.140625,
           "E":19.7109375,"F":18.42578125,"G":9.140625}
def _f(s): return Font(**s)

def build_nota_sheet(wb,title,rows,nama,alamat,tanggal):
    ws=wb.create_sheet(title=title[:31])
    # Layout SELALU 5 kolom. GT TIDAK punya kolom sendiri: penanda GT (dan GL)
    # dimasukkan ke kolom "No. GUD" mengikuti xlsx 1 (lihat loop data di bawah).
    has_gt=any(str(rr.get("gt") or "").strip().lower()=="gt" for rr in rows)
    c_gud,c_brt,c_net,c_hrg,c_jml = 1,2,3,4,5
    headers=["No. GUD","BRUTO","NETTO","HARGA","JUMLAH"]
    widths=[COL_WIDTH["A"],COL_WIDTH["C"],COL_WIDTH["D"],COL_WIDTH["E"],COL_WIDTH["F"],COL_WIDTH["G"]]
    for i,w in enumerate(widths,start=1):
        ws.column_dimensions[get_column_letter(i)].width=w
    L_net=get_column_letter(c_net); L_hrg=get_column_letter(c_hrg); L_jml=get_column_letter(c_jml)
    for r,h in {2:19.5,3:22.5,7:15.75,8:16.5}.items(): ws.row_dimensions[r].height=h
    # Sisipkan Logo Tembakau di A2 jika file logo.png tersedia
    try:
        logo_candidates = ["logo.png", os.path.join(os.path.dirname(__file__), "logo.png"), os.path.join(os.path.dirname(__file__), "..", "logo.png")]
        for lp in logo_candidates:
            if os.path.exists(lp):
                from openpyxl.drawing.image import Image as XLImage
                img = XLImage(lp)
                img.width = 48
                img.height = 48
                ws.add_image(img, "A2")
                break
    except Exception:
        pass
    ws["A2"].font=_f(FONT_TITLE); ws["A2"].alignment=Alignment(horizontal="left",vertical="center")
    ws["B2"]="NOTA PEMBELIAN TEMBAKAU 2026"
    ws["B2"].font=_f(FONT_TITLE); ws["B2"].alignment=Alignment(vertical="center")
    for r,(lab,val) in {4:("Nama    :",nama),5:("Alamat    :",alamat),6:("Tgl/Hr/Thn  :",tanggal)}.items():
        a=ws.cell(r,1,lab); a.font=_f(FONT_LABEL); a.alignment=Alignment(horizontal="right")
        bb=ws.cell(r,2,val if val else None); bb.font=_f(FONT_VALUE)
        if r==4:
            bb.number_format="[$-421]dd\\ mmmm\\ yyyy;@"
            d4=ws.cell(4,4); d4.font=_f(FONT_VALUE)
            d4.alignment=Alignment(horizontal="center")
            d4.number_format='[$-F800]dddd", "mmmm\\ dd", "yyyy'
    for i,h in enumerate(headers,start=1):
        c=ws.cell(8,i,h); c.font=_f(FONT_HEAD); c.alignment=Alignment(horizontal="center")
        c.border=Border(top=MED,bottom=MED,left=MED,right=MED)
    first=9; last=first+len(rows)-1
    for idx,row in enumerate(rows):
        r=first+idx; ws.row_dimensions[r].height=15.75
        top_ef=NONE if idx==0 else THIN
        # No. GUD mengikuti penanda xlsx 1: "GT <no>" bila kolom GT terisi,
        # "GL <no>" bila kolom GL(A)="gl", selain itu NOMOR saja (tanpa awalan).
        if row.get("bs"):
            label_a = row.get("label","BS")
        elif str(row.get("gt") or "").strip().lower()=="gt":
            label_a = f"GT {row['no']}"
        elif str(row.get("gl") or "").strip().lower()=="gl":
            label_a = f"GL {row['no']}"
        else:
            label_a = row["no"]
        a=ws.cell(r,c_gud,label_a); a.font=_f(FONT_GL)
        a.alignment=Alignment(horizontal="center",vertical="center")
        a.border=Border(top=THIN,bottom=THIN,left=THIN,right=THIN)
        c=ws.cell(r,c_brt,row["bruto"]); c.font=_f(FONT_CELL)
        c.alignment=Alignment(horizontal="center",vertical="center")
        c.border=Border(top=THIN,bottom=THIN,left=THIN,right=THIN)
        d=ws.cell(r,c_net,row["netto"]); d.font=_f(FONT_CELL); d.alignment=Alignment(horizontal="center")
        d.border=Border(top=NONE,bottom=THIN,left=THIN,right=THIN)
        e=ws.cell(r,c_hrg,row["harga"]); e.font=_f(FONT_CELL); e.alignment=Alignment(horizontal="center")
        e.number_format="#,##0"; e.border=Border(top=top_ef,bottom=THIN,left=THIN,right=THIN)
        f=ws.cell(r,c_jml,f"={L_net}{r}*{L_hrg}{r}"); f.font=_f(FONT_CELL); f.alignment=Alignment(horizontal="center")
        f.number_format="#,##0"; f.border=Border(top=top_ef,bottom=THIN,left=THIN,right=THIN)
    sum_formula = f"=SUM({L_jml}{first}:{L_jml}{last})"
    # baris GL terakhir (untuk COUNTIF GT) = sebelum baris BS (kalau ada baris BS di akhir)
    n_bs=sum(1 for rr in rows if rr.get("bs"))
    last_gl=last-n_bs if n_bs else last
    box=Border(top=THIN,bottom=THIN,left=THIN,right=THIN)
    fr=last+1; r_jml=fr; r_pph=fr+1
    if has_gt:
        r_gt=fr+2; r_koli=fr+3; r_tot=fr+4
        total_formula=f"={L_jml}{r_jml}-{L_jml}{r_koli}-{L_jml}{r_gt}-{L_jml}{r_pph}"
    else:
        r_gt=None; r_koli=fr+2; r_tot=fr+3
        total_formula=f"={L_jml}{r_jml}-{L_jml}{r_koli}-{L_jml}{r_pph}"
    footer=[("JUMLAH ",sum_formula,"#,##0",False),
            ("PPH 0,5%",f"=CEILING({L_jml}{r_jml}*0.005,5000)","#,##0",False)]
    if has_gt:
        footer.append(("GT",f'=65000*COUNTIF(A{first}:A{last_gl},"GT*")','"Rp"#,##0',False))
    footer.append(("Koli",f"=COUNTA(A{first}:A{last})*5000",'"Rp"#,##0',False))
    footer.append(("TOTAL",total_formula,'"Rp"#,##0',True))
    for i,(lab,formula,nf,bold) in enumerate(footer):
        r=fr+i; ws.row_dimensions[r].height=15.75
        e=ws.cell(r,c_hrg,lab); e.font=_f(FONT_CELL)
        e.alignment=Alignment(horizontal="right"); e.border=box
        f=ws.cell(r,c_jml,formula); f.font=_f(FONT_TOTAL if bold else FONT_CELL)
        f.alignment=Alignment(horizontal="center"); f.number_format=nf; f.border=box
    ws.sheet_view.showGridLines=False
    ws.page_setup.orientation="portrait"
    ws.page_margins.left=1.062992125984252
    ws.page_margins.right=0.1968503937007874
    ws.page_margins.top=0.1968503937007874
    ws.page_margins.bottom=0.1968503937007874
    return ws

@dataclass
class Nota:
    no_start:int; no_end:int; title:Optional[str]=None; nama:Optional[str]=None
    tanggal:Optional[str]=None; alamat:Optional[str]=None

OUT_DIR = "nota"

def default_out_name(notas):
    parts="_".join(f"{n.no_start}-{n.no_end}" for n in notas)
    return f"nota_{parts}.xlsx"

_ILLEGAL_SHEET = re.compile(r"[\\/?*\[\]:]")
def _safe_title(s):
    s = _ILLEGAL_SHEET.sub("-", str(s)).strip().strip("'").strip()
    return (s or "NOTA")[:31]

def build_workbook(src_path,out_path,notas):
    if out_path is None: out_path=default_out_name(notas)
    if not os.path.dirname(out_path):
        os.makedirs(OUT_DIR, exist_ok=True)
        out_path=os.path.join(OUT_DIR, out_path)
    else:
        d=os.path.dirname(out_path)
        if d: os.makedirs(d, exist_ok=True)
    src=openpyxl.load_workbook(src_path,data_only=True)
    ws1=src["Buku Soter"] if "Buku Soter" in src.sheetnames else src.active
    wb=openpyxl.Workbook(); wb.remove(wb.active)
    used=set()
    for n in notas:
        rows=read_rows(ws1,n.no_start,n.no_end)
        nama_auto,tanggal_auto,alamat_auto=detect_info(ws1,n.no_start,n.no_end)
        nama=n.nama if n.nama is not None else nama_auto
        tanggal=n.tanggal if n.tanggal is not None else tanggal_auto
        alamat=n.alamat if n.alamat is not None else alamat_auto
        base=_safe_title(n.title or nama or f"GL {n.no_start}-{n.no_end}")
        title=base; k=2
        while title[:31] in used:
            suffix=f" ({k})"; title=base[:31-len(suffix)]+suffix; k+=1
        used.add(title[:31])
        build_nota_sheet(wb,title,rows,nama,alamat,tanggal)
    wb.save(out_path); return out_path

if __name__ == "__main__":
    import sys
    args = sys.argv[1:]
    if args:
        src = args[0]
        rest = args[1:]
        out = None
        if "-o" in rest:
            i = rest.index("-o")
            out = rest[i+1]
            rest = rest[:i] + rest[i+2:]
        notas = []
        for arg in rest:
            a, b = arg.replace("–", "-").split("-")
            notas.append(Nota(int(a), int(b)))
        if not notas:
            print("Pemakaian: python nota_generator.py <file_xlsx1> [-o Hasil.xlsx] 142-147 [148-163 ...]")
            sys.exit(1)
        out = build_workbook(src, out, notas)
        print(f"Selesai -> {out}  ({len(notas)} nota)")
    else:
        SRC = "Buku_Soter_1-1000 GREEND.xlsx"
        OUT = None
        DAFTAR_NOTA = [
            Nota(1, 35),
        ]
        out = build_workbook(SRC, OUT, DAFTAR_NOTA)
        print(f"Selesai -> {out}  ({len(DAFTAR_NOTA)} nota)")
